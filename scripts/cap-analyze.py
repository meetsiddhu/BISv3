#!/usr/bin/env python3
"""CAP Application Analyser - generates the stakeholder analysis workbook (cap-analyze skill).

Reads live CDS + CSV + manifest/launchpad files (never gen/ or node_modules/) and writes a
multi-sheet .xlsx: exec / architecture / SME / test viewpoints + full catalog sheets.
Usage: python3 scripts/cap-analyze.py [repo_root] [out_dir]
"""
import json, re, csv, sys, subprocess, datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(sys.argv[1] if len(sys.argv) > 1 else '.').resolve()
OUT_DIR = Path(sys.argv[2]) if len(sys.argv) > 2 else ROOT / 'docs' / 'cap-analysis'
EXCLUDE = {'gen', 'node_modules', 'mta_archives', 'client-package', 'dist', '.git', 'resources'}
SME = 'NEEDS_SME_REVIEW'

def walk(base, suffix):
    out = []
    for p in sorted((ROOT / base).rglob('*' + suffix)):
        if not any(part in EXCLUDE for part in p.parts):
            out.append(p)
    return out

def rel(p): return str(Path(p).resolve().relative_to(ROOT))

# ---------------- Phase 0: fingerprint ----------------
pkg = json.loads((ROOT / 'package.json').read_text())
mta = (ROOT / 'mta.yaml').read_text()
version = re.search(r'^version:\s*([\d.]+)', mta, re.M).group(1)
def sh(cmd):
    try: return subprocess.run(cmd, shell=True, cwd=ROOT, capture_output=True, text=True, timeout=30).stdout.strip()
    except Exception: return ''
git_branch, git_sha = sh('git rev-parse --abbrev-ref HEAD'), sh('git rev-parse --short HEAD')
xsec = json.loads((ROOT / 'xs-security.json').read_text())
scopes = [s['name'].replace('$XSAPPNAME.', '') for s in xsec.get('scopes', [])]
role_templates = {r['name']: [x.replace('$XSAPPNAME.', '') for x in r.get('scope-references', [])] for r in xsec.get('role-templates', [])}

STD_TERMS = ['AS 5100', 'AS5100', 'NHVR', 'Austroads', 'AGBM', 'BIMM', 'TfNSW', 'RMS', 'HVNL',
             'ISO 55000', 'ISO 55001', 'ISO 27001', 'SOC 2', 'GDPR', 'IRAP', 'Essential 8', 'ONRSR', 'GDA2020', 'EPSG']
std_hits = {}
for p in walk('db', '.cds') + walk('srv', '.cds') + walk('docs', '.md') + [ROOT / 'CLAUDE.md']:
    try: txt = p.read_text(errors='ignore')
    except Exception: continue
    for t in STD_TERMS:
        if t in txt: std_hits.setdefault(t, set()).add(rel(p))

# ---------------- Phase 1: CDS entities ----------------
def parse_entities(path):
    """Brace-counting parse of `entity X [: aspects] { fields }` blocks with line numbers."""
    ents, lines = [], path.read_text().splitlines()
    i = 0
    while i < len(lines):
        m = re.match(r'\s*entity\s+([\w.]+)\s*(?::\s*([\w,\s]+?))?\s*\{', lines[i])
        if m and 'projection on' not in lines[i]:
            name, aspects, start = m.group(1), (m.group(2) or '').strip(), i
            depth, body = lines[i].count('{') - lines[i].count('}'), []
            i += 1
            while i < len(lines) and depth > 0:
                depth += lines[i].count('{') - lines[i].count('}')
                if depth > 0: body.append((i + 1, lines[i]))
                i += 1
            fields, assoc, pending_ann = [], [], []
            for ln, raw in body:
                s = raw.strip()
                if s.startswith('@') or s.startswith('//'):
                    pending_ann.append(s); continue
                fm = re.match(r'(?:key\s+)?(\w+)\s*:\s*(.+?);', s)
                if fm:
                    fname, ftype = fm.group(1), fm.group(2)
                    ann = ' '.join(pending_ann); pending_ann = []
                    inline_c = raw.split('//', 1)[1].strip() if '//' in raw else ''
                    am = re.search(r'Association to (?:many\s+)?([\w.]+)', ftype)
                    if am: assoc.append((fname, am.group(1).split('.')[-1]))
                    fields.append({'name': fname, 'type': ftype.split(' ')[0].rstrip(';'),
                                   'required': 'not null' in ftype or 'key ' in s,
                                   'key': s.startswith('key '), 'ann': ann, 'comment': inline_c,
                                   'enum': bool(re.search(r'@assert\.range|enum', ftype + ann)),
                                   'line': ln})
            ents.append({'name': name.split('.')[-1], 'aspects': aspects, 'fields': fields,
                         'assoc': assoc, 'source': f'{rel(path)}:{start + 1}'})
        else:
            i += 1
    return ents

entities = []
for f in [ROOT / 'db' / 'schema.cds'] + walk('db/schema', '.cds') + [ROOT / 'db' / 'attributes-schema.cds']:
    if f.exists(): entities.extend(parse_entities(f))
ent_by_name = {e['name']: e for e in entities}

# ---------------- Phase 1: services / projections / actions ----------------
def parse_services(path):
    txt = path.read_text(); out = []
    for sm in re.finditer(r'(@requires\s*:\s*(\[[^\]]*\]|[\'\"\w]+)\s*)?service\s+(\w+)', txt):
        svc = sm.group(3)
        seg = txt[sm.end():]
        nxt = re.search(r'\nservice\s+\w+', seg)
        if nxt: seg = seg[:nxt.start()]
        requires = re.findall(r"'(\w+)'", sm.group(2) or '')
        projs = []
        for pm in re.finditer(r'entity\s+(\w+)\s+as\s+projection\s+on\s+([\w.]+)', seg):
            pre = seg[:pm.start()][-700:]
            ro = '@readonly' in pre.rsplit(';', 1)[-1]
            grants = re.findall(r"grant\s*:\s*(\[[^\]]*\]|'[\w]+')\s*,\s*to\s*:\s*(\[[^\]]*\]|'[\w]+')", pre.rsplit(';', 1)[-1])
            projs.append({'name': pm.group(1), 'on': pm.group(2).split('.')[-1], 'readonly': ro,
                          'grants': [(re.sub(r"[\[\]' ]", '', g), re.sub(r"[\[\]' ]", '', t)) for g, t in grants]})
        acts = [{'kind': am.group(1), 'name': am.group(2)} for am in re.finditer(r'\b(action|function)\s+(\w+)\s*\(', seg)]
        out.append({'service': svc, 'requires': requires, 'projections': projs, 'actions': acts, 'source': rel(path)})
    return out

services = []
for f in sorted(set(walk('srv', '.cds'))):
    if '_i18n' in str(f) or f.name in ('common.cds',): continue
    services.extend(parse_services(f))

# audit coverage per service implementation file
audit_js = {}
for f in walk('srv', '.js'):
    t = f.read_text(errors='ignore')
    audit_js[rel(f)] = ('writeChangeLogs' in t or 'logAudit' in t)

# ---------------- Phase 1: tiles / inbounds / controllers ----------------
flp = json.loads((ROOT / 'app' / 'appconfig' / 'fioriSandboxConfig.json').read_text())
groups = flp['services']['LaunchPage']['adapter']['config']['groups']
inbounds = flp['services']['ClientSideTargetResolution']['adapter']['config']['inbounds']
def inbound_for(target):
    key = target.lstrip('#')
    for k, v in inbounds.items():
        if k.replace('-', '-') == key or (v['semanticObject'] + '-' + v['action']) == key: return k, v
    return None, None
controllers = [rel(p) for p in walk('app', '.controller.js') if 'webapp' in str(p)]

# curated tile metadata (traceable: app folder per inbound component)
TILE_OUTCOME = {
    'Dashboard': ('Portfolio insight at a glance', 'Time-to-answer for portfolio questions', 'Exec / Asset Manager'),
    'Bridges': ('Single governed bridge register', 'Register completeness %', 'Asset Manager'),
    'Restrictions': ('Statutory restriction lifecycle + audit', 'Active restrictions managed; time-to-publish', 'Operations / Compliance'),
    'MapView': ('Geographic context for decisions', SME, 'Engineer / Planner'),
    'Prioritisation': ('Defensible, reproducible funding ranking', 'Top-decile $ + band distribution; audit reproducibility', 'Exec / Engineer / Auditor'),
    'BridgeInspections': ('Inspection evidence captured to TfNSW levels', 'Overdue-inspection count', 'Inspector / Engineer'),
    'BridgeDefects': ('Engineering defects tracked, linked to EAM', 'Open defects by severity', 'Engineer'),
    'BridgeCapacities': ('Load/capacity (AS 5100) per structure', 'Structures with current load rating %', 'Structural Engineer'),
    'MassUpload': ('Bulk governed data intake', 'Rows loaded / rejected', 'Data Admin'),
    'MassEdit': ('Controlled bulk corrections', SME, 'Data Admin'),
    'BmsAdmin': ('Config-driven methodology governance', 'Config version currency', 'BMS Admin'),
    'AttributeClasses': ('Extensible attribute dictionary', SME, 'BMS Admin'),
    'EAMMapping': ('BIS-EAM value mapping (clean core)', 'Mapping coverage %', 'Integration Admin'),
    'BridgeRisk': ('Operational risk worklist', 'Fleet risk distribution', 'Asset Manager'),
    'NetworkPortfolio': ('Portfolio analytics by network', SME, 'Planner / Exec'),
    'RestrictionsDashboard': ('Restriction posture overview', 'Active/scheduled restriction counts', 'Operations'),
    'ChangeDocuments': ('Full audit trail visibility', 'Audit coverage of CUD events', 'Auditor'),
}

# ---------------- Phase 1: rules / seeds / tests ----------------
rules = []
for f in walk('srv', '.js'):
    for i, line in enumerate(f.read_text(errors='ignore').splitlines(), 1):
        m = re.search(r'req\.(error|reject)\(\s*(\d{3})?\s*,?\s*[\'"`](.{5,90}?)[\'"`]', line)
        if m: rules.append({'file': rel(f), 'line': i, 'verb': m.group(1), 'status': m.group(2) or '400', 'msg': m.group(3)})

seeds = []
for f in sorted((ROOT / 'db' / 'data').glob('*.csv')):
    try: n = max(0, sum(1 for _ in f.open()) - 1)
    except Exception: n = 0
    seeds.append({'file': f.name, 'rows': n, 'entity': f.stem.split('-')[-1]})

LOOKUP_FILES = [s for s in seeds if s['entity'] in (
    'ElementTypes', 'ImportanceLevels', 'Networks', 'LaneAvailabilityTypes', 'EAMCodeMapping')]
lookup_rows = []
for s in LOOKUP_FILES:
    with (ROOT / 'db' / 'data' / s['file']).open() as fh:
        rd = csv.DictReader(fh)
        for r in rd:
            cols = list(r.keys())
            lookup_rows.append({'category': s['entity'], 'key': r.get(cols[0], ''),
                                'text': r.get(cols[1], '') if len(cols) > 1 else '',
                                'extra': r.get(cols[2], '') if len(cols) > 2 else '',
                                'source': 'db/data/' + s['file']})

tests = []
for f in sorted((ROOT / 'test').glob('*.test.js')):
    t = f.read_text(errors='ignore')
    tests.append({'file': f.name, 'tests': len(re.findall(r'\b(?:test|it)\s*\(', t)),
                  'describes': len(re.findall(r'\bdescribe\s*\(', t))})

# ---------------- Phase 3 helpers ----------------
SUGGEST = [
    (r'conditionRating|conditionState', 'AS 5100.7 Table 4 (condition states) / TfNSW 1-5'),
    (r'postingStatus|loadLimit|loadRating|ratingStandard', 'AS 5100.7 + NHVR Bridge Assessment Guidelines'),
    (r'restrictionType|restrictionUnit|speedLimit|massLimit', 'NHVR Load Restrictions Schedule 2018 / HVNL s147'),
    (r'inspection(Type|Level|Date)', 'Austroads AGBM Part 6 / TfNSW inspection levels 1-3'),
    (r'defect(Severity|Extent|Type)', 'Austroads BIMM 6.2'),
    (r'gvm|gcm|permit', 'HVNL s155 / NHVR permit manual'),
    (r'structureType|material', 'Austroads AGBT04-14 2'),
    (r'scour', 'Austroads AGBT10 5 (scour)'),
    (r'hml|pbs|freightRoute|network', 'NHVR PBS network classification'),
    (r'gazette', 'State gazette under HVNL'),
    (r'latitude|longitude|geoJson|crs', 'GDA2020 (EPSG:7844)'),
    (r'owner|custodian|lga', 'TfNSW (NSW) asset custodianship'),
]
def suggest(field):
    for pat, std in SUGGEST:
        if re.search(pat, field, re.I): return std
    return ''
def authoritative(f):
    blob = f['ann'] + ' ' + f['comment']
    hits = [t for t in STD_TERMS if t in blob]
    return '; '.join(hits)

# ---------------- Phase 4: workbook ----------------
OUT_DIR.mkdir(parents=True, exist_ok=True)
today = datetime.date.today().isoformat()
out_path = OUT_DIR / f"{pkg['name']}_v{version}_{today}.xlsx"
wb = Workbook(); wb.remove(wb.active)
HDR_FILL = PatternFill('solid', fgColor='1565C0')
def sheet(name, headers, rows, widths=None):
    ws = wb.create_sheet(name[:31])
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF'); c.fill = HDR_FILL
        c.alignment = Alignment(vertical='center')
    for r in rows: ws.append([('' if v is None else v) for v in r])
    ws.freeze_panes = 'A2'
    for idx, h in enumerate(headers, 1):
        w = (widths or {}).get(h)
        if not w:
            w = min(60, max(len(str(h)) + 2, max((len(str(r[idx-1])) for r in rows if idx-1 < len(r)), default=10) + 2))
        ws.column_dimensions[get_column_letter(idx)].width = w
    return ws

# T1 Tiles
t1 = []
for g in groups:
    for t in g['tiles']:
        k, v = inbound_for(t['properties'].get('targetURL', ''))
        comp = (v or {}).get('resolutionResult', {}).get('additionalInformation', '')
        t1.append([t['id'], g['id'], t['properties']['title'], t['properties'].get('icon', ''),
                   t['properties'].get('targetURL', ''), comp.replace('SAPUI5.Component=', ''),
                   'view+ (XSUAA-gated in services)', 'app/appconfig/fioriSandboxConfig.json'])
sheet('T1 Tiles', ['tile_id', 'group', 'title', 'icon', 'route', 'component', 'roles_visible', 'source'], t1)

# T2 Tile functionalities
t2 = []
for g in groups:
    for t in g['tiles']:
        tid = t['id']
        o = TILE_OUTCOME.get(tid, (SME, SME, SME))
        k, v = inbound_for(t['properties'].get('targetURL', ''))
        comp = (v or {}).get('resolutionResult', {}).get('additionalInformation', '').replace('SAPUI5.Component=', '')
        t2.append([tid, o[0], comp, 'OData V4 CRUD per @restrict', 'srv/services + srv/*.js',
                   'C/R/U/D per security model (02b)', 'app/appconfig/fioriSandboxConfig.json'])
sheet('T2 Tile Functionalities', ['tile', 'functionality', 'component', 'entity_actions_used', 'related_handler', 'crud_summary', 'source'], t2)

# T3 Field catalog (top entities by relevance + size)
TOP = ['Bridges', 'BridgeRestrictions', 'PrioritisationAssessment', 'PrioritisationConfig',
       'EamWorkRequest', 'BridgeInspections', 'BridgeDefects', 'BridgeCapacities', 'BridgeElements']
t3 = []
for en in TOP:
    e = ent_by_name.get(en)
    if not e: continue
    for f in e['fields']:
        purpose = f['comment'] or re.sub(r'@title\s*:\s*', '', f['ann'])[:80] or SME
        t3.append([en, f['name'], f['type'], 'Y' if f['required'] else '', 'Y' if f['enum'] else '',
                   purpose, authoritative(f), suggest(f['name']),
                   e['source'].split(':')[0] + ':' + str(f['line'])])
sheet('T3 Field Catalog', ['entity', 'field', 'type', 'required', 'enum_ref', 'purpose', 'standard_ref', 'suggested_standard_ref', 'source'], t3, {'purpose': 45, 'suggested_standard_ref': 42})

# T4 Configurations
t4 = []
for e in entities:
    if re.search(r'Config|RiskBand|AssetClassStrategy|Rubric', e['name']):
        seed = next((s['rows'] for s in seeds if s['entity'] == e['name']), 0)
        t4.append([e['name'], seed, (e['fields'][1]['comment'] if len(e['fields']) > 1 and e['fields'][1]['comment'] else SME),
                   'admin (XSUAA)', 'BMS Administration', e['source']])
sheet('T4 Configurations', ['config_entity', 'seed_rows', 'what_it_configures', 'editable_by_role', 'admin_ui', 'source'], t4)

# T5 Lookups
sheet('T5 Lookups', ['category', 'key', 'text', 'extra', 'source'],
      [[l['category'], l['key'], l['text'], l['extra'], l['source']] for l in lookup_rows])

# T6 Data model + ER
t6 = [[e['name'], e['source'].split(':')[0], len(e['fields']),
       next((f['name'] for f in e['fields'] if f['key']), 'ID(cuid)' if 'cuid' in e['aspects'] else ''),
       ', '.join(f'{a[0]}->{a[1]}' for a in e['assoc'][:6]), e['aspects'], e['source']] for e in entities]
ws6 = sheet('T6 Data Model', ['entity', 'file', 'fields', 'key', 'associations', 'aspects', 'source'], t6)
core = [e for e in entities if e['assoc'] or any(a[1] == e['name'] for x in entities for a in x['assoc'])]
core_names = {e['name'] for e in core}
mer = ['erDiagram']
for e in core[:28]:
    for a in e['assoc']:
        if a[1] in core_names: mer.append(f'  {e["name"]} }}o--|| {a[1]} : "{a[0]}"')
mermaid_txt = '\n'.join(dict.fromkeys(mer))
er_png, er_ok = OUT_DIR / 'er-diagram.png', False
try:
    mmd = OUT_DIR / 'er.mmd'; mmd.write_text(mermaid_txt)
    subprocess.run(f'mmdc -i "{mmd}" -o "{er_png}" -b white --width 1600', shell=True, timeout=120, capture_output=True)
    er_ok = er_png.exists()
except Exception: pass
row0 = len(t6) + 4
ws6.cell(row=row0, column=1, value='Mermaid ER (editable):').font = Font(bold=True)
ws6.cell(row=row0 + 1, column=1, value=mermaid_txt)
if er_ok:
    try:
        from openpyxl.drawing.image import Image as XImage
        img = XImage(str(er_png)); img.anchor = f'A{row0 + 3}'; ws6.add_image(img)
    except Exception: er_ok = False

# 01 Exec summary
flags = [['prioritisationEnabled', 'default ON (SystemConfig kill-switch)']]
exec_rows = [
    ['Application', f"{pkg['name']} (BIS — Bridge Information System)"],
    ['Version (MTA)', version], ['Git', f'{git_branch} @ {git_sha}'],
    ['Stack', f"CAP {pkg['dependencies'].get('@sap/cds', '')} · Node 20 · HANA Cloud · Fiori/UI5 · XSUAA"],
    ['Tiles / apps', f'{sum(len(g["tiles"]) for g in groups)} tiles · {len(set(c.split("/")[1] for c in controllers))} UI apps'],
    ['Entities / services', f'{len(entities)} entities · {len(services)} services'],
    ['Tests', f'{sum(t["tests"] for t in tests)} tests across {len(tests)} suites'],
    ['Compliance lexicon found', ', '.join(sorted(std_hits.keys())) or 'none'],
    ['Feature flags', '; '.join(f'{k}: {v}' for k, v in flags)],
    ['Posture', 'Clean-core side-by-side complement to SAP EAM — EAM never modified'],
    ['Cost drivers', 'HANA Cloud HDI + CF runtime (512M srv dev / 1G HA prod via mtaext) + html5 repo'],
    ['Design reference', 'docs/SOLUTION-DESIGN-BRIDGE-LIFECYCLE.md'],
]
sheet('01 Exec Summary', ['item', 'value'], exec_rows, {'value': 90})

# 01a Business outcomes
sheet('01a Business Outcomes', ['tile', 'business_outcome', 'kpi', 'stakeholder'],
      [[tid, o[0], o[1], o[2]] for tid, o in TILE_OUTCOME.items()])

# 01b Compliance register (authoritative only)
ent_std = {}
for e in entities:
    hits = set()
    for f in e['fields']:
        for t in STD_TERMS:
            if t in (f['ann'] + f['comment']): hits.add(t)
    if hits: ent_std[e['name']] = sorted(hits)
sheet('01b Compliance Register', ['entity', 'standards_referenced (authoritative, from CDS annotations/comments)'],
      [[k, ', '.join(v)] for k, v in sorted(ent_std.items())] or [['(none found in CDS annotations)', 'Standards lexicon present in docs: ' + ', '.join(sorted(std_hits.keys()))]])

# 02 Architecture
mods = re.findall(r'- name:\s*([\w-]+)\s*\n\s*type:\s*([\w.]+)', mta)
arch = [['Module: ' + m, t] for m, t in mods]
arch += [['Resource: ' + r, ''] for r in re.findall(r'resources:\n(?:.|\n)*?', '')]
arch += [['Memory (srv)', re.search(r'memory:\s*(\w+)', mta).group(1) if re.search(r'memory:\s*(\w+)', mta) else ''],
         ['HA', 'mtaext/bms-prod.mtaext → instances:3 / 1G'],
         ['Topology', 'Launchpad/approuter → CAP OData V4 → HANA HDI; Destination → S/4HANA EAM (read-only refs + work-request push Next); Later: Datasphere/BDC + SAC, Event Mesh/IoT'],
         ['Integrations', 'EAM value mapping (EAMCodeMapping/EAMFieldMapping), EamWorkRequest outbound queue, GIS (GDA2020)']]
sheet('02 Architecture', ['component', 'detail'], arch, {'detail': 80})

# 02a API catalog
api = []
for s in services:
    for p in s['projections']:
        crud = 'R' if p['readonly'] else ('|'.join(f'{g}->{t}' for g, t in p['grants']) or 'per service @requires')
        api.append([s['service'], p['name'], p['on'], crud, ','.join(s['requires']) or 'authenticated',
                    'Y' if audit_js.get(s['source'].replace('.cds', '.js'), False) else '', s['source']])
    for a in s['actions']:
        api.append([s['service'], f"{a['kind']} {a['name']}()", '', 'action', ','.join(s['requires']) or '', '', s['source']])
sheet('02a API Catalog', ['service', 'entity_or_action', 'on_entity', 'crud/grants', 'requires_scope', 'audit', 'source'], api)

# 02b Security model
sec = [[rt, ', '.join(sc), xsec.get('xsappname', '')] for rt, sc in role_templates.items()]
sheet('02b Security Model', ['role_template', 'scopes', 'xsappname'], sec)

# 03 Business rules
sheet('03 Business Rules', ['rule_id', 'handler', 'line', 'verb', 'status', 'message'],
      [[f'BR-{i+1:03d}', r['file'], r['line'], r['verb'], r['status'], r['msg']] for i, r in enumerate(rules)],
      {'message': 70, 'handler': 40})

# 03a Role-capability matrix
cap_rows = []
for s in services:
    for p in s['projections']:
        for g, t in (p['grants'] or [('READ' if p['readonly'] else 'CRUD', ','.join(s['requires']) or 'any-auth')]):
            cap_rows.append([t, s['service'] + '.' + p['name'], g, s['source']])
sheet('03a Role-Capability', ['role/scope', 'entity', 'grant', 'source'], cap_rows)

# 04 Test & assurance
t4a = [[t['file'], t['describes'], t['tests']] for t in tests]
t4a.append(['TOTAL', sum(t['describes'] for t in tests), sum(t['tests'] for t in tests)])
ws = sheet('04 Test & Assurance', ['suite', 'describes', 'tests'], t4a)
ws.cell(row=len(t4a) + 3, column=1, value='Known gaps: no mutation-score report; perf baselines not present in repo; UI e2e covered via live smoke (manual+scripted).')

wb.save(out_path)

# ---------------- report ----------------
sme_count = sum(1 for sh_ in wb.worksheets for row in sh_.iter_rows() for c in row if c.value == SME)
print(json.dumps({
    'workbook': rel(out_path), 'sheets': [ws.title for ws in wb.worksheets],
    'rows_per_sheet': {ws.title: ws.max_row - 1 for ws in wb.worksheets},
    'needs_sme_review_cells': sme_count, 'er_rendered': er_ok,
    'entities': len(entities), 'services': len(services), 'rules': len(rules),
    'tests_total': sum(t['tests'] for t in tests)
}, indent=1))
